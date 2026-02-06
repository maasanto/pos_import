# Copyright (c) 2025, Dokos SAS and contributors
# For license information, please see license.txt

import csv
from datetime import datetime
from decimal import Decimal
from io import BytesIO, StringIO

from pos_import.pos_import.parsers.base import BasePOSParser, POSLine, POSPayment, POSReport


class RestomaxParser(BasePOSParser):
	"""
	Parser for Restomax export files (Excel or CSV).

	File format expectations:
	- Excel or CSV file with columns: N° Z, Date clôture, ID Restomax, Compte général, Description, TVA, DEBIT, CREDIT
	- ID Restomax column contains source codes for item/payment mapping
	- Lines are duplicated (need deduplication)
	- Account 700xxx range = revenues (CREDIT column)
	- Account 451xxx range = VAT collected
	- Account 580xxx range = payments (DEBIT column)

	IMPORTANT: Restomax export doubles all amounts, so we divide by 2 to get correct values
	"""

	REVENUE_ACCOUNT_PREFIX = "700"
	VAT_ACCOUNT_PREFIX = "451"
	PAYMENT_ACCOUNT_PREFIX = "580"

	REQUIRED_COLUMNS = ["N° Z", "Date clôture", "Compte général", "DEBIT", "CREDIT", "ID Restomax"]

	def validate_file(self, file_content: bytes) -> tuple[bool, str]:
		"""Validate that the file is a valid Restomax export."""
		try:
			rows = self._read_file(file_content)
			if not rows:
				return False, "Fichier vide"

			headers = list(rows[0].keys())
			missing = [col for col in self.REQUIRED_COLUMNS if col not in headers]

			if missing:
				return False, f"Colonnes manquantes: {', '.join(missing)}"

			return True, ""
		except Exception as e:
			return False, f"Erreur de lecture du fichier: {e}"

	def parse(self, file_content: bytes) -> list[POSReport]:
		"""Parse the Restomax file and return a list of POS reports."""
		rows = self._read_file(file_content)

		reports_data: dict = {}
		seen_lines: set[tuple] = set()

		for row in rows:
			report_num = row.get("N° Z")
			if not report_num:
				continue

			report_num = str(report_num).strip()

			account = str(row.get("Compte général") or "").strip()
			id_restomax = str(row.get("ID Restomax") or "").strip()
			original_description = str(row.get("Description") or "").strip()
			# For revenue lines with empty description, use "Others" as fallback label
			description = original_description or "Others"
			debit = self._parse_number(row.get("DEBIT"))
			credit = self._parse_number(row.get("CREDIT"))

			line_key = (report_num, account, original_description, str(debit), str(credit))

			if line_key in seen_lines:
				continue
			seen_lines.add(line_key)

			if report_num not in reports_data:
				reports_data[report_num] = {
					"date": row.get("Date clôture"),
					"revenues": [],
					"payments": [],
					"vat": [],
				}

			tva_rate = self._parse_number(row.get("TVA"))

			if account.startswith(self.REVENUE_ACCOUNT_PREFIX):
				# Skip total/summary lines by keyword
				description_lower = original_description.lower()
				if any(
					keyword in description_lower
					for keyword in ["total", "sous-total", "subtotal", "ca global", "ca tvac", "ca hors"]
				):
					continue

				# Restomax doubles all amounts - divide by 2 to get correct values
				amount = (credit - debit) / 2
				if amount > 0:
					reports_data[report_num]["revenues"].append({
						"account": account,
						"id_restomax": id_restomax,
						"description": description,
						"amount": amount,
						"tva_rate": tva_rate,
					})

			elif account.startswith(self.VAT_ACCOUNT_PREFIX):
				# VAT collected (451000 accounts) - these are the actual VAT amounts
				if "total" in original_description.lower():
					continue

				# Skip lines without ID Restomax (summary lines)
				if not id_restomax:
					continue

				# Restomax doubles all amounts - divide by 2 to get correct values
				amount = (credit - debit) / 2
				if amount != 0:
					reports_data[report_num]["vat"].append({
						"account": account,
						"description": description,
						"amount": amount,
						"tva_rate": tva_rate,
					})

			elif account.startswith(self.PAYMENT_ACCOUNT_PREFIX):
				# Skip total lines
				if original_description.startswith("Total CA") or original_description.startswith("Total PAIEMENT"):
					continue

				# Restomax doubles all amounts - divide by 2 to get correct values
				amount = (debit - credit) / 2
				if amount > 0:
					reports_data[report_num]["payments"].append({
						"account": account,
						"id_restomax": id_restomax,
						"description": description,
						"amount": amount,
					})

		reports = []
		for report_num, data in reports_data.items():
			report = POSReport(
				report_number=report_num,
				report_date=self._parse_date(data["date"]),
			)

			# Parse actual VAT amounts from 451000 accounts
			for vat_data in data["vat"]:
				tva_rate = vat_data["tva_rate"]
				if tva_rate not in report.vat_by_rate:
					report.vat_by_rate[tva_rate] = Decimal(0)
				report.vat_by_rate[tva_rate] += vat_data["amount"]

			# Build line items with HT amounts only
			for line_data in data["revenues"]:
				tva_rate = line_data["tva_rate"]
				# The amounts in 700000 accounts are HT (net)
				net = line_data["amount"]

				# Use ID Restomax as source_code (for item mapping)
				# If empty, fall back to description for unmapped items
				source_code = line_data["id_restomax"] or line_data["description"]

				# We don't calculate tax here - use actual VAT from vat_by_rate
				report.lines.append(
					POSLine(
						source_code=source_code,
						description=line_data["description"],
						net_amount=net.quantize(Decimal("0.01")),
						tax_rate=tva_rate,
						tax_amount=Decimal(0),  # Not used - actual VAT in report.vat_by_rate
						gross_amount=net.quantize(Decimal("0.01")),  # Will be recalculated
					)
				)

			for payment_data in data["payments"]:
				# Use ID Restomax as source_code for payment mapping
				# If empty, fall back to description
				source_code = payment_data["id_restomax"] or payment_data["description"]
				report.payments.append(
					POSPayment(
						source_code=source_code,
						source_name=payment_data["description"],
						amount=payment_data["amount"].quantize(Decimal("0.01")),
					)
				)

			reports.append(report)

		return sorted(reports, key=lambda r: (r.report_date, r.report_number))

	def _read_file(self, file_content: bytes) -> list[dict]:
		"""Read file content as CSV or Excel and return list of dicts."""
		# Try CSV first (detect delimiter automatically)
		try:
			text = self._decode_content(file_content)
			first_line = text.split("\n")[0]

			# Detect delimiter (semicolon or comma)
			delimiter = ";" if ";" in first_line else ","
			reader = csv.DictReader(StringIO(text), delimiter=delimiter)
			rows = list(reader)
			if rows:  # Successfully read CSV
				return rows
		except Exception:
			pass

		# Try Excel
		try:
			import openpyxl
			wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
			ws = wb.active

			headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
			rows = []
			for row in ws.iter_rows(min_row=2, values_only=True):
				rows.append(dict(zip(headers, row)))
			return rows
		except Exception:
			pass

		raise ValueError("Impossible de lire le fichier (format CSV ou Excel attendu)")

	def _decode_content(self, content: bytes) -> str:
		"""Decode bytes to string, trying multiple encodings."""
		for encoding in ["utf-8-sig", "utf-8", "iso-8859-1", "cp1252"]:
			try:
				return content.decode(encoding)
			except UnicodeDecodeError:
				continue
		raise ValueError("Impossible de décoder le fichier")

	def _parse_number(self, value) -> Decimal:
		"""Parse European number format (comma decimal, space thousands separator)."""
		if value is None:
			return Decimal(0)

		if isinstance(value, (int, float)):
			return Decimal(str(value))

		if isinstance(value, Decimal):
			return value

		text = str(value).strip()
		if not text:
			return Decimal(0)

		# Remove non-breaking spaces and regular spaces (thousands separator)
		text = text.replace("\u00a0", "").replace("\u202f", "").replace(" ", "")
		# Replace comma with dot (decimal separator)
		text = text.replace(",", ".")

		try:
			return Decimal(text)
		except Exception:
			return Decimal(0)

	def _parse_date(self, date_value) -> datetime:
		"""Parse various date formats to a date object."""
		if isinstance(date_value, datetime):
			return date_value.date()

		if hasattr(date_value, "date"):
			return date_value.date()

		if isinstance(date_value, str):
			for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S"]:
				try:
					return datetime.strptime(date_value.strip(), fmt).date()
				except ValueError:
					continue

		return datetime.now().date()
